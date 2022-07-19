from openpyxl import workbook, load_workbook

wb = load_workbook('books.xlsx')
ws = wb.active
print(ws['A1'].value)

take_input1 = input(" write col A: ")
take_input2 = input(" write col B: ")
take_input3 = input(" write col C: ")
take_input4 = input(" write col D: ")

rows = (take_input1, take_input2, take_input3, take_input4)

ws.append(rows)

wb.save('books.xlsx')